import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from cryptography.fernet import Fernet
import io

# Set up Streamlit app title
st.title("Transaction Report Generator")

# Generate a key for encryption (persistent if shared securely)
# For demonstration, we create a new key each time, but save this in production
key = Fernet.generate_key()
cipher = Fernet(key)

# Encrypt file content
def encrypt_file(file_data):
    return cipher.encrypt(file_data)

# Decrypt file content
def decrypt_file(file_data):
    return cipher.decrypt(file_data)

# File uploader
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file:
    # Decrypt file content (if applicable)
    file_data = uploaded_file.read()
    try:
        # Attempt decryption
        decrypted_data = decrypt_file(file_data)
        file = io.BytesIO(decrypted_data)
        st.success("File decrypted successfully.")
    except:
        # Use original file data if decryption fails
        file = io.BytesIO(file_data)
        st.warning("Decryption failed or file wasn't encrypted. Proceeding with original file.")

    # Load the data into pandas
    data = pd.read_excel(file)
    st.write("Data Preview:", data.head())  # Show preview of the data

    # Processing: Example of Summation by Category
    if 'Category' in data.columns and 'Amount' in data.columns:
        summary = data.groupby('Category')['Amount'].sum().reset_index()
        st.write("Summary Report:", summary)

        # Export summary to Excel for download
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary.to_excel(writer, index=False, sheet_name='Summary')
            # Apply formatting if desired (e.g., bold headers)
            workbook = writer.book
            worksheet = writer.sheets['Summary']
            for col_num, value in enumerate(summary.columns.values, 1):
                col_letter = get_column_letter(col_num)
                cell = worksheet[f"{col_letter}1"]
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        output.seek(0)  # Reset pointer to beginning

        # Optionally encrypt the report before download
        encrypted_output = encrypt_file(output.read())
        encrypted_io = io.BytesIO(encrypted_output)
        encrypted_io.seek(0)

        # Download button
        st.download_button(
            label="Download Encrypted Report",
            data=encrypted_io,
            file_name="encrypted_weekly_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("The uploaded file must contain 'Category' and 'Amount' columns.")
